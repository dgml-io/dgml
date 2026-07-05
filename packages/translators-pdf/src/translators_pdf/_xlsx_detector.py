# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Active-data "island" detection in an Excel sheet.

An active cell is non-empty, non-whitespace, and not in a hidden row/column.
Cells are grouped by BFS with a connection distance set by ``row_gap`` /
``col_gap`` so that visually separate tables become separate islands (pages).
"""

from __future__ import annotations

from typing import Any

from openpyxl.utils import get_column_letter


def is_row_hidden(sheet: Any, r: int) -> bool:
    if r in sheet.row_dimensions:
        return bool(sheet.row_dimensions[r].hidden)
    return False


def is_col_hidden(sheet: Any, c: int) -> bool:
    col_letter = get_column_letter(c)
    if col_letter in sheet.column_dimensions:
        return bool(sheet.column_dimensions[col_letter].hidden)
    return False


def find_islands(sheet: Any, row_gap: int = 2, col_gap: int = 2) -> list[tuple[int, int, int, int]]:
    """Return island bounding boxes ``(min_row, min_col, max_row, max_col)``.

    Sorted top-to-bottom, then left-to-right. ``row_gap``/``col_gap`` are the
    maximum consecutive empty rows/columns tolerated within one island.
    """
    active_cells: list[tuple[int, int]] = []
    for r in range(1, sheet.max_row + 1):
        if is_row_hidden(sheet, r):
            continue
        for c in range(1, sheet.max_column + 1):
            if is_col_hidden(sheet, c):
                continue
            val = sheet.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                active_cells.append((r, c))

    if not active_cells:
        return []

    # An empty gap of G means a cell-to-cell distance of G+1 is still connected.
    max_r_dist = row_gap + 1
    max_c_dist = col_gap + 1

    visited: set[tuple[int, int]] = set()
    islands: list[tuple[int, int, int, int]] = []

    for r, c in active_cells:
        if (r, c) in visited:
            continue
        island_cells: list[tuple[int, int]] = []
        queue: list[tuple[int, int]] = [(r, c)]
        visited.add((r, c))
        while queue:
            curr_r, curr_c = queue.pop(0)
            island_cells.append((curr_r, curr_c))
            for nr, nc in active_cells:
                if (nr, nc) not in visited and (
                    abs(nr - curr_r) <= max_r_dist and abs(nc - curr_c) <= max_c_dist
                ):
                    visited.add((nr, nc))
                    queue.append((nr, nc))

        min_row = min(cr for cr, _ in island_cells)
        max_row = max(cr for cr, _ in island_cells)
        min_col = min(cc for _, cc in island_cells)
        max_col = max(cc for _, cc in island_cells)
        islands.append((min_row, min_col, max_row, max_col))

    islands.sort(key=lambda bbox: (bbox[0], bbox[1]))
    return islands
