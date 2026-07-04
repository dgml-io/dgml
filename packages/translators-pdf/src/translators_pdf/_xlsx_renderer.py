# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Render detected Excel islands into a styled PDF.

Each island becomes one page, preserving fonts, fills, borders, merged cells,
alignment, and common number formats. Output is written to a path; the xlsx
converter wraps this to return bytes.
"""

from __future__ import annotations

import datetime
import html
import os
from typing import Any, Literal

import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Flowable,
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    TableStyle,
)


class PageMetadataFlowable(Flowable):
    """Non-rendering flowable that passes page metadata to the canvas."""

    def __init__(
        self, filename: str, sheet_name: str, island_index: int, total_islands: int
    ) -> None:
        super().__init__()
        self.width = 0
        self.height = 0
        self.filename = filename
        self.sheet_name = sheet_name
        self.island_index = island_index
        self.total_islands = total_islands

    def draw(self) -> None:
        # The canvas has no static slots for these; NumberedCanvas reads them
        # back via getattr when drawing page furniture.
        self.canv.page_filename = self.filename  # type: ignore[attr-defined]
        self.canv.page_sheet_name = self.sheet_name  # type: ignore[attr-defined]
        self.canv.page_island_index = self.island_index  # type: ignore[attr-defined]
        self.canv.page_total_islands = self.total_islands  # type: ignore[attr-defined]


class NumberedCanvas(canvas.Canvas):
    """Two-pass canvas for running headers/footers and 'Page X of Y'."""

    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, **kwargs)
        self._saved_page_states: list[dict[str, Any]] = []

    def showPage(self) -> None:  # noqa: N802
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()  # type: ignore[attr-defined]  # reportlab Canvas internal

    def save(self) -> None:
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_elements(num_pages)
            super().showPage()
        super().save()

    def draw_page_elements(self, page_count: int) -> None:
        self.saveState()
        width, height = self._pagesize  # type: ignore[attr-defined]  # reportlab Canvas internal
        margin = 36

        filename = getattr(self, "page_filename", "Unknown")
        sheet_name = getattr(self, "page_sheet_name", "Sheet")
        island_index = getattr(self, "page_island_index", 0)
        total_islands = getattr(self, "page_total_islands", 0)

        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#1F4E79"))
        self.drawString(margin, height - 38, f"Workbook: {filename}")

        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#555555"))
        self.drawString(margin + 180, height - 38, f"|   Sheet: {sheet_name}")

        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#2C3E50"))
        self.drawRightString(
            width - margin, height - 38, f"Island {island_index} of {total_islands}"
        )

        self.setStrokeColor(colors.HexColor("#D5DBDB"))
        self.setLineWidth(0.5)
        self.line(margin, height - 44, width - margin, height - 44)

        self.line(margin, 46, width - margin, 46)
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#7F8C8D"))
        gen_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.drawString(margin, 30, f"Generated: {gen_time}   |   dgml xlsx converter")

        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#34495E"))
        self.drawRightString(
            width - margin,
            30,
            f"Page {self._pageNumber} of {page_count}",  # type: ignore[attr-defined]  # reportlab internal
        )
        self.restoreState()


def format_cell_value(val: Any, fmt: str | None) -> str:
    """Format a raw cell value using common Excel number-format conventions."""
    if val is None:
        return ""
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        if not fmt or fmt == "General":
            if isinstance(val, float):
                s = str(val)
                if "." in s and len(s.split(".")[1]) > 4:
                    return f"{val:.4f}".rstrip("0").rstrip(".")
            return str(val)
        if "%" in fmt:
            if "0.00" in fmt:
                return f"{val * 100:.2f}%"
            if "0.0" in fmt:
                return f"{val * 100:.1f}%"
            return f"{val * 100:.0f}%"
        if "$" in fmt:
            if "0.00" in fmt:
                return f"${val:,.2f}"
            if "0.0" in fmt:
                return f"${val:,.1f}"
            return f"${val:,.0f}"
        if "#,##0" in fmt:
            if "0.00" in fmt:
                return f"{val:,.2f}"
            if "0.0" in fmt:
                return f"{val:,.1f}"
            return f"{val:,.0f}"
        if "0.00" in fmt:
            return f"{val:.2f}"
        if "0.0" in fmt:
            return f"{val:.1f}"
    return str(val)


def hex_to_reportlab_color(rgb_hex: str | None, default_color: Any = colors.black) -> Any:
    """Convert an openpyxl ARGB hex string to a ReportLab color."""
    if not rgb_hex or not isinstance(rgb_hex, str):
        return default_color
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    if len(rgb_hex) == 6:
        try:
            return colors.HexColor(f"#{rgb_hex}")
        except ValueError:
            pass
    return default_color


def get_border_weight_and_color(side: Any) -> tuple[float, Any]:
    """Border thickness (points) and color from an openpyxl Side."""
    if not side or side.style is None:
        return 0, None
    style_map = {
        "thin": 0.5,
        "medium": 1.0,
        "thick": 1.8,
        "double": 1.5,
        "dashed": 0.5,
        "hair": 0.3,
    }
    weight = style_map.get(side.style, 0.5)
    color = colors.HexColor("#BDC3C7")
    if side.color and side.color.rgb and side.color.type == "rgb":
        color = hex_to_reportlab_color(side.color.rgb, default_color=color)
    return weight, color


def render_pdf(
    wb_path: str,
    sheet_islands: dict[str, list[tuple[int, int, int, int]]],
    pdf_path: str,
    orientation: str = "landscape",
) -> None:
    """Render all detected islands across all sheets into a single PDF at ``pdf_path``."""
    filename = os.path.basename(wb_path)
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    if orientation == "portrait":
        page_size = letter
        p_width, _p_height = letter
    else:
        page_size = landscape(letter)
        p_width, _p_height = landscape(letter)

    doc = SimpleDocTemplate(
        pdf_path, pagesize=page_size, leftMargin=36, rightMargin=36, topMargin=54, bottomMargin=54
    )
    printable_width = p_width - 72
    story: list[Any] = []
    total_islands = sum(len(islands) for islands in sheet_islands.values())
    island_counter = 0

    for sheet_name, islands in sheet_islands.items():
        if not islands:
            continue
        sheet = wb[sheet_name]

        for idx, (min_r, min_c, max_r, max_c) in enumerate(islands):
            island_counter += 1
            story.append(PageMetadataFlowable(filename, sheet_name, idx + 1, len(islands)))

            col_widths: list[float] = []
            for c in range(min_c, max_c + 1):
                col_letter = get_column_letter(c)
                dim = (
                    sheet.column_dimensions.get(col_letter)
                    if col_letter in sheet.column_dimensions
                    else None
                )
                char_width = dim.width if dim and dim.width else 12.0
                if char_width < 1.0:
                    char_width = 12.0
                col_widths.append(char_width * 7.0)

            total_raw_width = sum(col_widths)
            if total_raw_width > printable_width:
                scale_factor = printable_width / total_raw_width
                scaled_col_widths = [w * scale_factor for w in col_widths]
            else:
                scaled_col_widths = col_widths

            avg_col_width = sum(scaled_col_widths) / len(scaled_col_widths)
            if avg_col_width < 25.0:
                font_size = 5.0
            elif avg_col_width < 35.0:
                font_size = 6.0
            elif avg_col_width < 45.0:
                font_size = 7.0
            else:
                font_size = 8.5
            leading = font_size + 2.0

            grid_data: list[list[Any]] = []
            table_styles: list[Any] = [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3.0),
                ("TOPPADDING", (0, 0), (-1, -1), 3.0),
                ("LEFTPADDING", (0, 0), (-1, -1), 4.0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4.0),
            ]

            for rng in sheet.merged_cells.ranges:
                if (
                    rng.min_col <= max_c
                    and rng.max_col >= min_c
                    and rng.min_row <= max_r
                    and rng.max_row >= min_r
                ):
                    rel_start_c = max(rng.min_col, min_c) - min_c
                    rel_start_r = max(rng.min_row, min_r) - min_r
                    rel_end_c = min(rng.max_col, max_c) - min_c
                    rel_end_r = min(rng.max_row, max_r) - min_r
                    if rel_start_c <= rel_end_c and rel_start_r <= rel_end_r:
                        table_styles.append(
                            ("SPAN", (rel_start_c, rel_start_r), (rel_end_c, rel_end_r))
                        )

            for r_idx, r in enumerate(range(min_r, max_r + 1)):
                row_data: list[Any] = []
                for c_idx, c in enumerate(range(min_c, max_c + 1)):
                    cell = sheet.cell(row=r, column=c)
                    val = cell.value
                    formatted_val = format_cell_value(val, cell.number_format)
                    escaped_val = html.escape(formatted_val)

                    is_bold = bool(cell.font and cell.font.bold)
                    is_italic = bool(cell.font and cell.font.italic)

                    text_color_hex = "#333333"
                    if (
                        cell.font
                        and cell.font.color
                        and cell.font.color.rgb
                        and cell.font.color.type == "rgb"
                    ):
                        text_color_hex = cell.font.color.rgb
                        if len(text_color_hex) == 8:
                            text_color_hex = text_color_hex[2:]
                        text_color_hex = f"#{text_color_hex}"

                    align_code: Literal[0, 1, 2, 4] = 0
                    if cell.alignment and cell.alignment.horizontal:
                        align_str = cell.alignment.horizontal.lower()
                        if align_str == "center":
                            align_code = 1
                        elif align_str == "right":
                            align_code = 2
                        elif align_str == "justify":
                            align_code = 4
                    elif isinstance(val, bool):
                        align_code = 1
                    elif isinstance(val, (int, float)):
                        align_code = 2

                    style = ParagraphStyle(
                        name=f"Style_{sheet_name}_{island_counter}_{r}_{c}",
                        fontName="Helvetica-Bold" if is_bold else "Helvetica",
                        fontSize=font_size,
                        leading=leading,
                        textColor=colors.HexColor(text_color_hex),
                        alignment=align_code,
                    )
                    p_text = escaped_val
                    if is_bold:
                        p_text = f"<b>{p_text}</b>"
                    if is_italic:
                        p_text = f"<i>{p_text}</i>"
                    row_data.append(Paragraph(p_text, style))

                    if cell.fill and cell.fill.fill_type == "solid":
                        bg_color = hex_to_reportlab_color(
                            cell.fill.start_color.rgb, default_color=None
                        )
                        if bg_color:
                            table_styles.append(
                                ("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), bg_color)
                            )

                    if cell.border:
                        for side, op in (
                            (cell.border.top, "LINEABOVE"),
                            (cell.border.bottom, "LINEBELOW"),
                            (cell.border.left, "LINEBEFORE"),
                            (cell.border.right, "LINEAFTER"),
                        ):
                            weight, color = get_border_weight_and_color(side)
                            if weight > 0 and color:
                                table_styles.append(
                                    (op, (c_idx, r_idx), (c_idx, r_idx), weight, color)
                                )

                grid_data.append(row_data)

            # LongTable (vs Table) splits a tall island across pages instead of
            # demanding it fit one frame — so large sheets render rather than
            # failing. (A single row taller than a page still can't be split;
            # that surfaces as the clear error below.)
            table = LongTable(grid_data, colWidths=scaled_col_widths)
            table.setStyle(TableStyle(table_styles))
            table.hAlign = "CENTER"
            story.append(table)

            if island_counter < total_islands:
                story.append(PageBreak())

    try:
        doc.build(story, canvasmaker=NumberedCanvas)
    except Exception as exc:
        # When an island still doesn't fit (e.g. a single row taller than a
        # page, or a sheet so wide that column-scaling forces text to wrap into
        # an unsplittable row), reportlab raises — and its own "flowable too
        # large" reporter can crash with an opaque `'>' not supported between
        # NoneType and float`. Translate any build failure into an actionable
        # message instead of leaking that.
        raise RuntimeError(
            "could not lay the workbook out as PDF — a data island is too large or too "
            "wide to fit on a page (a single row taller than a page cannot be split). "
            f"Underlying error: {exc}"
        ) from exc
